import time

import pyautogui
from PIL import Image
import openpyxl
import imageio
from tkinter import messagebox, filedialog
import os

from openpyxl.workbook import Workbook

# Diccionario de coherencia
coherencia_dict = {
    "SENSE ADEQUACIÓ": "CENTRE (SENSE TRASLLAT)",
    "KIT PROJECCIÓ": "CENTRE (SENSE TRASLLAT)",
    "PANEL INTERACTIU": "CENTRE (SENSE TRASLLAT)",
    "PISSARRA": "CENTRE (SENSE TRASLLAT)",
    "PROYECTOR": "CENTRE (SENSE TRASLLAT)",
    "ALTRES": "CENTRE (SENSE TRASLLAT)"
}
# Diccionario de mensajes
mensaje_dict = {
    "SENSE ADEQUACIÓ": "Resum del material a retirar: No es fa retirada perquè es monta suport amb rodes.",
    "OTROS": "La destinació de tots el elements a retirar és CENTRE (SENSE TRASLLAT)."
}


class ImageConverterService:
    def convertir_imagenes_en_directorio(self, directorio, formato_salida, barra_progreso, ventana):
        try:
            archivos = self._obtener_archivos_validos(directorio)
            total_archivos = len(archivos)
            barra_progreso["maximum"] = total_archivos
            errores = []

            for indice, archivo in enumerate(archivos, start=1):
                try:
                    self._convertir_imagen(archivo, formato_salida)
                except Exception as e:
                    errores.append(f"{archivo}")
                finally:
                    barra_progreso["value"] = indice
                    ventana.update_idletasks()

            fotos_convertidas = total_archivos - len(errores)
            if errores:
                errores_mensaje = "\n".join(errores)
                messagebox.showerror(
                    "Errores de conversión",
                    f"Se han convertido {fotos_convertidas} de {total_archivos} imágenes."
                    f"\nLas siguientes imágenes no se han podido convertir debido a un problema de corrupción de datos:"
                    f"\n{errores_mensaje}\nProceda a recorte manual.")
            else:
                messagebox.showinfo("Completado",
                                    f"Se han convertido {fotos_convertidas} de {total_archivos} imágenes exitosamente.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al convertir imágenes: {e}")

    def _obtener_archivos_validos(self, directorio):
        archivos_validos = []
        for raiz, _, archivos in os.walk(directorio):
            for archivo in archivos:
                if archivo.lower().endswith(('.png', '.jpg', '.jpeg', '.heic')):
                    archivos_validos.append(os.path.join(raiz, archivo))
        return archivos_validos

    def _convertir_imagen(self, ruta_archivo, formato_salida):
        nombre_archivo, extension_archivo = os.path.splitext(ruta_archivo)
        ruta_salida = f"{nombre_archivo}.{formato_salida.lower()}"

        try:
            if extension_archivo.lower() == '.heic':
                try:
                    num_fotogramas = imageio.get_reader(ruta_archivo).get_length()
                    if num_fotogramas == 0:
                        raise ValueError("La imagen HEIC no contiene ningún fotograma.")
                    imagen = imageio.imread(ruta_archivo)
                except Exception as e:
                    raise ValueError(f"Error al leer la imagen HEIC: {e}")

                imagen = Image.fromarray(imagen)
                imagen = imagen.convert("RGB")
            else:
                imagen = Image.open(ruta_archivo)
                imagen = imagen.convert("RGB")

            imagen.save(ruta_salida, formato_salida.upper())

            if ruta_archivo != ruta_salida:
                os.remove(ruta_archivo)

        except Exception as e:
            raise RuntimeError(f"Error al convertir la imagen {ruta_archivo}: {e}")


class ExcelService:
    def completar_check(self, ruta_archivo, celdas, estado_ok):
        try:
            libro = openpyxl.load_workbook(ruta_archivo)

            # Completar las hojas con estado_ok
            for nombre_hoja in libro.sheetnames[1:]:
                hoja = libro[nombre_hoja]
                for celda in celdas:
                    hoja[celda] = estado_ok

            # Guardar el libro de Excel
            libro.save(ruta_archivo)
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar el archivo de Excel: {e}")

    def firmar_acta(self, ruta_archivo, celdas):
        try:
            libro = openpyxl.load_workbook(ruta_archivo)
            datos_instalador = {
                "DNI": "45489700R",
                "NOM": "David Quiñonero",
                #"CÀRREC": "Tècnic",
                #"TELÈFON": "647999633"
                "FECHA": time.strftime("%d/%m/%Y")
            }

            for celda, valor in datos_instalador.items():
                libro.active[celdas[celda]] = valor

            # Guardar el libro de Excel
            libro.save(ruta_archivo)
        except Exception as e:
            messagebox.showerror("Error", f"Error al firmar el documento: {e}")
            return False

        return True

    def encontrar_celdas_para_firmar_en_acta(self, ruta_archivo):
        palabras_clave = ["DNI", "NOM", "CÀRREC", "TELÈFON"]
        resultado = {}
        try:
            # Cargar el archivo Excel
            libro = openpyxl.load_workbook(ruta_archivo)
            hoja = libro.active

            # Recorrer las filas de la columna F
            for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=6, max_col=6):
                for celda in fila:
                    if celda.value in palabras_clave:
                        # Obtener la referencia de la celda en la columna G de la misma fila
                        referencia_celda = f"G{celda.row}"
                        resultado[celda.value] = referencia_celda
            if resultado["NOM"]:
                fecha = int(resultado["NOM"][1:]) + 8
                resultado["FECHA"] = f"G{fecha}"
                print(resultado["FECHA"])
        except Exception as e:
            print(f"Ha ocurrido un error: {e}")
        return resultado

    def comentario_automatico(self, ruta_archivo, replanteo=False):
        # Cargar el archivo Excel
        wb = openpyxl.load_workbook(ruta_archivo)

        # Si replanteo es False, agregar los textos específicos en la primera hoja
        if not replanteo:
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value == "OBSERVACIONS":
                        ws.cell(row=cell.row + 1, column=1, value="Instal·lació de punt de xarxa:")
                        ws.cell(row=cell.row + 2, column=1, value="Instal·lació de punt elèctric:")
                        break

        # Iterar sobre las hojas a partir de la segunda
        for sheet_name in wb.sheetnames[1:]:
            ws = wb[sheet_name]

            # Buscar "TIPUS" en la columna A
            for row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value == "TIPUS":
                        tipus_fila = cell.row
                        tipus_valor = ws.cell(row=tipus_fila + 1, column=1).value
                        if replanteo:
                            desti_valor = ws.cell(row=tipus_fila + 1, column=3).value
                        else:
                            desti_valor = ws.cell(row=tipus_fila + 1, column=2).value

                        # Verificar la coherencia
                        if tipus_valor in coherencia_dict and coherencia_dict[tipus_valor] == desti_valor:
                            coherente = True
                            mensaje = mensaje_dict.get(tipus_valor, mensaje_dict["OTROS"])
                        else:
                            coherente = False
                            mensaje = "Tipus no es correspon amb destinació, revisar a INDIC"

                        # Buscar "OBSERVACIONS" en la columna A y poner el mensaje
                        for obs_row in ws.iter_rows(min_col=1, max_col=1, min_row=1, max_row=ws.max_row):
                            for obs_cell in obs_row:
                                if obs_cell.value == "OBSERVACIONS":
                                    if coherente and tipus_valor == "SENSE ADEQUACIÓ":
                                        ws.cell(row=obs_cell.row + 4, column=1, value=mensaje)
                                    elif coherente:
                                        ws.cell(row=obs_cell.row + 5, column=1, value=mensaje_dict["OTROS"])
                                    else:
                                        ws.cell(row=obs_cell.row + 5, column=1, value=mensaje)
                                    break
                        break

        # Guardar los cambios en el archivo Excel
        wb.save(ruta_archivo)
        return True


class ExtractorService:

    def solicitar_archivo(self):

        archivo = filedialog.askopenfilename(
            title="Selecciona el archivo Excel",
            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
        )

        return archivo

    # Función para solicitar la ruta de destino
    def solicitar_destino(self):

        destino = filedialog.askdirectory(
            title="Selecciona la carpeta de destino",
        )
        if destino:
            return destino
        else:
            return False

    # Función para extraer la información de las celdas B2 y C4 de todas las hojas
    def extraer_informacion(self, ruta_archivo, tipo):
        libro = openpyxl.load_workbook(ruta_archivo, data_only=True)
        datos = []
        hoja1 = libro.worksheets[0]
        print("tipo: ", tipo)

        if tipo == "instalación":
            codigo_centro = hoja1['C13'].value  # Extraer el título de la primera hoja
            nombre_centro = hoja1['F13'].value
            print(codigo_centro, nombre_centro)
            for hoja in libro.worksheets[1:]:
                C16 = hoja['C19'].value
                K17 = hoja['M19'].value
                SN = hoja['G24'].value
                SACE = hoja['K24'].value
                print(C16, K17, SN, SACE)
                datos.append((C16, K17, SN, SACE))

            return datos, nombre_centro, codigo_centro, tipo
        if tipo == "replanteo":
            codigo_centro = hoja1['C9'].value  # Extraer el título de la primera hoja
            nombre_centro = hoja1['F9'].value

            for hoja in libro.worksheets[1:]:
                C16 = hoja['C16'].value
                K17 = hoja['K17'].value
                datos.append((C16, K17))

            return datos, nombre_centro, codigo_centro, tipo

    # Función para crear un nuevo archivo Excel con la información extraída
    def crear_nuevo_archivo(self, datos, nombre_centro, codigo_centro, ruta_destino, tipo):
        nuevo_libro = Workbook()
        nueva_hoja = nuevo_libro.active
        nueva_hoja.title = "Datos Extraídos"

        if nombre_centro and tipo == "instalación":
            # Escribir los títulos
            nueva_hoja['A2'] = "Nombre"
            nueva_hoja['B2'] = "Aula"
            nueva_hoja['C2'] = "Alias"
            nueva_hoja['D2'] = "SN"
            nueva_hoja['E2'] = "SACE"
            # Escribir los datos en las celdas B1 y C1
            for index, (b2, c4, SN, SACE) in enumerate(datos, start=1):
                if b2 and c4:
                    nueva_hoja[f'A{index}'] = b2
                    nueva_hoja[f'B{index}'] = b2[10:]
                    nueva_hoja[f'C{index}'] = c4
                    nueva_hoja[f'D{index}'] = SN
                    nueva_hoja[f'E{index}'] = SACE

        elif nombre_centro and tipo == "replanteo":
            # Escribir los títulos
            nueva_hoja['A1'] = "Nombre"
            nueva_hoja['B1'] = "Aula"
            for index, (b2, c4) in enumerate(datos, start=1):
                if b2 and c4:
                    nueva_hoja[f'A{index}'] = b2
                    nueva_hoja[f'B{index}'] = c4
        else:
            messagebox.showwarning("Advertencia", "No se encontraron datos para extraer.")
            return False

        nombre_archivo = ""
        if tipo == "instalación":
            nombre_archivo = os.path.join(ruta_destino, f"{codigo_centro}_{nombre_centro}_relación_aula_instalación.xlsx")
        if tipo == "replanteo":
            nombre_archivo = os.path.join(ruta_destino, f"{codigo_centro}_{nombre_centro}_relación_aula_replanteo.xlsx")
        if nombre_archivo == "":
            messagebox.showwarning("Advertencia", "Error Inesperado.")
            return False
        nuevo_libro.save(nombre_archivo)
        messagebox.showinfo("Éxito", f"Nuevo archivo creado como {nombre_archivo}.")

        # Abrir el archivo recién creado
        try:
            os.startfile(nombre_archivo)
        except AttributeError:
            # Para sistemas no Windows
            os.system(f"open {nombre_archivo}")


class AutomationService:
    def __init__(self):
        self.columnas_info = {
            'codigo_centro': 'int',
            'nombre_centro': 'String',
            'codigo_nombre_centro': 'String',
            'nombre_aula': 'String',
            'numero_edificio': [1, 2, 3, 4, 5],
            'planta': [1, 2, 3, 4, 5, -1, -2, -3],
            'etapa': ['INFANTIL', 'PRIMARIA', 'ESO', 'BATXILLERAT', 'CICLOS FORMATIVOS', 'SENSE DEFINIR'],
            'edificio_con_modulos': ['SI (Obligatori suport de carro de rodes)', 'No'],
            'panel': ['65"', '75"', '86"'],
            'tipo_soporte': [
                'Paret FIX', 'Paret Regulable', 'Rodes FIX', 'Rodes Regulable',
                'Paret amb Potes FIX (ESPECIAL PARET DE PLADUR O PARETS AMB POCA ESTABILITAT)',
                'Paret amb Potes Regulable (ESPECIAL PARET DE PLADUR O PARETS AMB POCA ESTABILITAT)'
            ],
            'el_soporte_es_diferente': ['NO', 'SI'],
            'tipo_pared': ['Totxo', 'Formigó', 'Mòdul', 'Pladur', 'Pedra', 'Rajola', 'Sense definir'],
            'presa_electrica': [
                'NO ÉS NECESSARI PRESA ELÈCTRICA', 'SI, 1m', 'SI, 3m', 'SI, 5m',
                'SI, 10m', 'SI, més de 10m'
            ],
            'red': [
                'NO ES REQUEREIX. La distancia entre punt de xarxa i lateral del panell es inferior a 2m.',
                'NO ES REQUEREIX. La conexió serà per wifi.', 'SI, 5m', 'SI, 7m',
                'SI, 10m', 'SI, 15m', 'SI, 20m', 'SI, més de 20m'
            ]
        }
        self.datos = []

    def automatizar_teclas(self):
        ruta_archivo = filedialog.askopenfilename(title="Selecciona un archivo Excel", filetypes=[("Archivos Excel", "*.xlsx")])
        if ruta_archivo:
            try:
                wb = openpyxl.load_workbook(ruta_archivo)
                ws = wb.active

                self.datos = []
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    fila = {
                        'codigo_centro': row[0].value,
                        'nombre_centro': row[1].value,
                        'codigo_nombre_centro': row[2].value,
                        'nombre_aula': row[3].value,
                        'numero_edificio': row[4].value,
                        'planta': row[5].value,
                        'etapa': row[6].value,
                        'edificio_con_modulos': row[7].value,
                        'panel': row[8].value,
                        'tipo_soporte': row[9].value,
                        'el_soporte_es_diferente': row[10].value,
                        'tipo_pared': row[11].value,
                        'presa_electrica': row[12].value,
                        'red': row[13].value
                    }
                    self.datos.append(fila)

                messagebox.showinfo("Información", "Datos recogidos correctamente. Ahora puedes ver los detalles.")
                return self.datos
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar el archivo Excel: {e}")
        else:
            messagebox.showwarning("Advertencia", "Por favor, seleccione un archivo Excel.")

    def procesar_datos(self, dato):
        """
        Lógica para procesar los datos recogidos y ejecutar una macro en función de los datos.
        """
        # Ejecutar la secuencia

        pyautogui.press('tab')

        # Pestaña ubicación
        self.ubicacion(dato)
        self.equipament(dato)
        self.finalizar()

        # Pasar de ventana
        messagebox.showinfo(f"{dato['nombre_aula']} creada", f"Aula creada correctamente.")

    def ubicacion(self, dato):

        # Numero edificio
        if int(dato['numero_edificio']) in self.columnas_info['numero_edificio']:
            pyautogui.write(f'{dato["numero_edificio"]}')

        # Siguiente
        pyautogui.press('tab')

        # Lógica edificio con modulos
        if dato['edificio_con_modulos'].lower().strip() == "no":
            pyautogui.press('n')
        else:
            pyautogui.press('s')

        # Siguiente
        pyautogui.press('tab')

        # Planta
        if int(dato['planta']) >= 0:
            pyautogui.write(f'{dato["planta"]}')
        elif int(dato['planta']) == -1:
            pyautogui.press('-')
        elif int(dato['planta']) == -2:
            pyautogui.press('-')
            pyautogui.press('-')
        elif int(dato['planta']) == -3:
            pyautogui.press('-')
            pyautogui.press('-')
            pyautogui.press('-')

        pyautogui.press('tab')

        # Etapa
        pyautogui.press(f'{dato["etapa"][0]}')

        pyautogui.press('tab')

        # Nombre aula
        pyautogui.write(f'{dato['nombre_aula']}')

        # Pasar de ventana
        pyautogui.press('tab')
        pyautogui.press('enter')
        pyautogui.press('tab')


    def equipament(self, dato):
        metros_de_red = None
        amb_rodes = False

        # Panel
        pyautogui.write(f'{dato["panel"][11]}')

        pyautogui.press('tab')

        # Tipo soporte
        if dato['tipo_soporte']:
            # Paret FIX
            if dato['tipo_soporte'] == 'Paret FIX' or dato['tipo_soporte'] == 'Suport de paret fixe - LP4390F-B':
                pyautogui.press('p')
                pyautogui.press('p')
            # Paret Regulable
            elif dato['tipo_soporte'] == 'Paret Regulable' or dato['tipo_soporte'] == 'Suport de paret regulable - LPHA7090':
                pyautogui.press('r')
            # Rodes FIX
            elif dato['tipo_soporte'] == 'SUPORT AMB RODES FIXE - FS20200M-B':
                pyautogui.press('a')
                amb_rodes = True
            # Rodes Regulable
            elif dato['tipo_soporte'] == 'Suport amb rodes regulable - FS20404HM-B':
                pyautogui.press('r')
                pyautogui.press('r')
                amb_rodes = True
            else:
                pyautogui.press('p')

        pyautogui.press('tab')

        # Pared
        if dato['tipo_pared']:
            # Totxo
            if dato['tipo_pared'][:2].lower() == 'to':
                pyautogui.press('t')
            # Formigó
            elif dato['tipo_pared'][:2].lower() == 'fo':
                pyautogui.press('f')
            # Pladur
            elif dato['tipo_pared'][:2].lower() == 'pl':
                pyautogui.press('p')
                pyautogui.press('p')
            # Piedra
            elif dato['tipo_pared'][:2].lower() == 'pe':
                pyautogui.press('p')
            # Mòdul
            elif dato['tipo_pared'][:2].lower() == 'mo' or 'mò':
                pyautogui.press('m')
            # Rajola
            elif dato['tipo_pared'][:2].lower() == 'ra':
                pyautogui.press('r')
            else:

                pyautogui.press('s')

        pyautogui.press('tab')

        # Fuetó
        if dato['red'][:2].strip().lower() == 'no':
            pyautogui.press('n')
        elif dato['red'][:2].strip().lower() == 'si':
            metros_de_red = StringService.solo_numeros(dato['red'])
            if dato['red'][3:].strip().lower() == '1m':
                pyautogui.press('f')
            elif dato['red'][3:].strip().lower() == '3m':
                pyautogui.press('f')
                pyautogui.press('f')
                pyautogui.press('f')
            elif dato['red'][3:].strip().lower() == '5m':
                pyautogui.press('f')
                pyautogui.press('f')
                pyautogui.press('f')
                pyautogui.press('f')
            else:
                pyautogui.press('f')
                pyautogui.press('f')

        pyautogui.press('tab')

        if dato['presa_electrica'][:2].strip().lower() == 'si':
            pyautogui.press('space')
            metros_electrica = StringService.solo_numeros(dato['presa_electrica'])
            pyautogui.press('tab')
            pyautogui.write(f'{metros_electrica}')

        pyautogui.press('tab')

        # Colocar metros de red
        if metros_de_red:
            pyautogui.press('space')
            pyautogui.press('tab')
            pyautogui.write(f'{metros_de_red}')

        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')

        if amb_rodes:
            pyautogui.press('space')
            pyautogui.press('tab')
            pyautogui.press('p')

        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('space')

    def finalizar(self):

        pyautogui.press('tab')

        # Sense Adequació
        pyautogui.press('s')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')

        # Centre sense trasllat
        pyautogui.press('c')

        #siguiente ventana
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('space')

        #finalizar
        # siguiente ventana
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('tab')
        pyautogui.press('space')

        #siguiente ventana
        pyautogui.press('tab')
        pyautogui.press('tab')


class StringService:

    @staticmethod
    def solo_numeros(texto_o_lista):
        texto = texto_o_lista
        return ''.join([c for c in texto if c.isdigit()])
